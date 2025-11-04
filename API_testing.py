#!/usr/bin/env python3
"""
Simplified MiniMax Clinical Evaluation System
For simplified Excel structure: Patient Text + 5 Treatment Plans
"""

import anthropic
import openpyxl
import json
import time
import pandas as pd
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
from datetime import datetime
import os

@dataclass
class MiniMaxEvaluation:
    """Data structure for MiniMax evaluation results"""
    case_id: int = 0
    model_name: str = ""
    patient_text: str = ""
    treatment_plan: str = ""
    rating: str = ""
    score_justification: str = ""
    key_observations: str = ""
    overall_assessment: str = ""
    clinical_reasoning: str = ""
    safety_considerations: str = ""
    evidence_quality: str = ""
    implementation_feasibility: str = ""
    specialist_notes: str = ""
    thinking_process: str = ""
    evaluation_timestamp: str = ""

class SimplifiedMiniMaxEvaluator:
    def __init__(self, api_key: str = None):
        """
        Initialize Simplified MiniMax Evaluator
        
        Args:
            api_key: MiniMax API key
        """
        self.api_key = api_key
        self.client = None
        if api_key:
            self._initialize_client()
    
    def _initialize_client(self):
        """Initialize MiniMax client with Anthropic SDK"""
        try:
            self.client = anthropic.Anthropic(
                base_url="https://api.minimax.io/anthropic", 
                api_key=self.api_key
            )
            print("✅ MiniMax client initialized successfully")
        except Exception as e:
            print(f"❌ Failed to initialize MiniMax client: {e}")
            raise
    
    def set_api_key(self, api_key: str):
        """Set or update API key"""
        self.api_key = api_key
        self._initialize_client()
    
    def _create_evaluation_prompt(self, patient_text: str, treatment_plan: str, model_name: str) -> str:
        """Create evaluation prompt for MiniMax"""
        return f"""You are a Board-Certified Sleep Medicine Specialist with 20+ years of clinical experience. 

**Patient Case:**
{patient_text}

**Treatment Plan to Evaluate ({model_name}):**
{treatment_plan}

**Evaluation Task:**
As an expert sleep medicine specialist, evaluate this treatment plan using these criteria:

## Rating Scale (A-E)
- **A (Excellent)**: Outstanding, highly clinically reliable with clear reasoning, appropriate risk awareness, and practical implementation
- **B (Very Good)**: Strong with good reasoning, minor improvement areas  
- **C (Good)**: Solid, mostly accurate, some advice could be more specific
- **D (Fair)**: Basic with notable gaps in safety, specificity, or appropriateness
- **E (Poor)**: Inadequate with significant concerns, inappropriate suggestions, or lack of clinical relevance

## Key Assessment Areas
- **Clinical Appropriateness**: Is this suitable for this specific patient?
- **Safety Awareness**: Are potential risks and contraindications addressed?
- **Evidence Quality**: Is the reasoning sound and evidence-based?
- **Implementation Feasibility**: Can this realistically be implemented?
- **Specialist Insights**: Additional sleep medicine expertise

**Provide your evaluation in this exact JSON format:**

{{
    "rating": "[A/B/C/D/E]",
    "score_justification": "Detailed explanation with clinical reasoning",
    "key_observations": "Specific observations about safety, appropriateness, clarity",
    "overall_assessment": "Summary of recommendation quality and clinical utility",
    "clinical_reasoning": "Analysis of clinical appropriateness for this case",
    "safety_considerations": "Evaluation of safety aspects and risk management",
    "evidence_quality": "Assessment of evidence-based reasoning quality",
    "implementation_feasibility": "Evaluation of real-world implementation potential",
    "specialist_notes": "Additional expert commentary from sleep medicine perspective"
}}

**Important**: 
- Output only valid JSON
- Focus on sleep medicine expertise
- Consider evidence-based practices
"""
    
    def evaluate_treatment_plan(self, 
                               patient_text: str, 
                               treatment_plan: str, 
                               model_name: str, 
                               case_id: int = 0) -> MiniMaxEvaluation:
        """
        Evaluate a treatment plan using MiniMax API
        
        Args:
            patient_text: Patient case description
            treatment_plan: AI-generated treatment plan to evaluate
            model_name: Name of the AI model that generated the plan
            case_id: Case identifier
            
        Returns:
            MiniMaxEvaluation object with detailed results
        """
        if not self.client:
            raise ValueError("MiniMax client not initialized. Please set API key first.")
        
        # Create evaluation prompt
        prompt = self._create_evaluation_prompt(patient_text, treatment_plan, model_name)
        
        try:
            print(f"🔍 Evaluating {model_name} case {case_id}...")
            
            # Make API call to MiniMax
            message = self.client.messages.create(
                model="MiniMax-M2",
                max_tokens=2000,
                system="You are a Board-Certified Sleep Medicine Specialist with 20+ years of clinical experience.",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ]
            )
            
            # Process response
            evaluation_data = {}
            thinking_content = ""
            
            for block in message.content:
                if block.type == "thinking":
                    thinking_content = block.thinking
                elif block.type == "text":
                    # Try to extract JSON from text response
                    text_content = block.text.strip()
                    
                    # Find JSON content
                    if text_content.startswith('{') and text_content.endswith('}'):
                        try:
                            evaluation_data = json.loads(text_content)
                        except json.JSONDecodeError:
                            # If direct parsing fails, try to find JSON in the text
                            import re
                            json_match = re.search(r'\{.*\}', text_content, re.DOTALL)
                            if json_match:
                                try:
                                    evaluation_data = json.loads(json_match.group())
                                except json.JSONDecodeError:
                                    print(f"⚠️ Could not parse JSON from response: {text_content[:200]}...")
                                    evaluation_data = {}
            
            # Create evaluation object
            evaluation = MiniMaxEvaluation(
                case_id=case_id,
                model_name=model_name,
                patient_text=patient_text,
                treatment_plan=treatment_plan,
                rating=evaluation_data.get('rating', 'C'),
                score_justification=evaluation_data.get('score_justification', ''),
                key_observations=evaluation_data.get('key_observations', ''),
                overall_assessment=evaluation_data.get('overall_assessment', ''),
                clinical_reasoning=evaluation_data.get('clinical_reasoning', ''),
                safety_considerations=evaluation_data.get('safety_considerations', ''),
                evidence_quality=evaluation_data.get('evidence_quality', ''),
                implementation_feasibility=evaluation_data.get('implementation_feasibility', ''),
                specialist_notes=evaluation_data.get('specialist_notes', ''),
                thinking_process=thinking_content,
                evaluation_timestamp=datetime.now().isoformat()
            )
            
            print(f"✅ Completed evaluation: {evaluation.rating} for {model_name}")
            return evaluation
            
        except Exception as e:
            print(f"❌ Error evaluating case {case_id}: {e}")
            # Return fallback evaluation
            return MiniMaxEvaluation(
                case_id=case_id,
                model_name=model_name,
                patient_text=patient_text,
                treatment_plan=treatment_plan,
                rating='C',
                score_justification=f"Evaluation failed due to: {str(e)}",
                key_observations="Technical error prevented evaluation",
                overall_assessment="Evaluation incomplete",
                thinking_process=f"Error: {str(e)}",
                evaluation_timestamp=datetime.now().isoformat()
            )
    
    def load_simplified_excel_data(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Load data from simplified Excel structure:
        Column 1: Patient Text
        Column 2: Treatment_Plan Claude
        Column 3: Treatment_Plan Deepseek (or Deepseek)
        Column 4: Treatment_Plan GPT
        Column 5: Treatment_Plan Grok
        Column 6: Treatment_Plan GPT5
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            cases = []
            
            # Start from row 3 (assuming row 1 is empty, row 2 has headers)
            for row_num in range(3, sheet.max_row + 1):
                case_data = self._extract_simplified_case_data(sheet, row_num)
                if case_data and case_data.get('patient_text'):
                    cases.append(case_data)
            
            print(f"📊 Loaded {len(cases)} cases from simplified Excel")
            return cases
            
        except Exception as e:
            print(f"❌ Error loading Excel data: {e}")
            return []
    
    def _extract_simplified_case_data(self, sheet, row_num: int) -> Dict[str, Any]:
        """Extract case data from simplified Excel structure"""
        case_data = {
            'case_id': row_num - 2,  # Account for header rows
            'patient_text': sheet.cell(row=row_num, column=1).value or '',
            'treatment_plans': {}
        }
        
        # Extract treatment plans based on simplified structure
        # Column 2: Treatment_Plan Claude
        claude_plan = sheet.cell(row=row_num, column=2).value
        if claude_plan:
            case_data['treatment_plans']['Claude'] = str(claude_plan)
        
        # Column 3: Treatment_Plan Deepseek (or just Deepseek)
        deepseek_plan = sheet.cell(row=row_num, column=3).value
        if deepseek_plan:
            case_data['treatment_plans']['Deepseek'] = str(deepseek_plan)
        
        # Column 4: Treatment_Plan GPT
        gpt_plan = sheet.cell(row=row_num, column=4).value
        if gpt_plan:
            case_data['treatment_plans']['ChatGPT'] = str(gpt_plan)
        
        # Column 5: Treatment_Plan Grok
        grok_plan = sheet.cell(row=row_num, column=5).value
        if grok_plan:
            case_data['treatment_plans']['Grok'] = str(grok_plan)
        
        # Column 6: Treatment_Plan GPT5
        gpt5_plan = sheet.cell(row=row_num, column=6).value
        if gpt5_plan:
            case_data['treatment_plans']['GPT-5'] = str(gpt5_plan)
        
        return case_data
    
    def batch_evaluate_simplified_excel(self, 
                                      excel_file_path: str, 
                                      output_file: str = "minimax_simplified_results.xlsx",
                                      start_row: int = 3,
                                      end_row: int = None) -> List[MiniMaxEvaluation]:
        """
        Batch evaluate all treatment plans from simplified Excel file
        """
        if not self.client:
            raise ValueError("MiniMax client not initialized. Please set API key first.")
        
        # Load Excel data
        cases = self.load_simplified_excel_data(excel_file_path)
        if not cases:
            print("❌ No cases loaded from Excel file")
            return []
        
        # Filter cases by row range
        if end_row is not None:
            cases = [case for case in cases if start_row + case['case_id'] - 3 <= end_row]
        
        all_evaluations = []
        
        for case in cases:
            case_id = case['case_id']
            patient_text = case['patient_text']
            
            for model_name, treatment_plan in case['treatment_plans'].items():
                print(f"\n🔍 Evaluating Case {case_id} - {model_name}")
                
                # Evaluate treatment plan
                evaluation = self.evaluate_treatment_plan(
                    patient_text=patient_text,
                    treatment_plan=treatment_plan,
                    model_name=model_name,
                    case_id=case_id
                )
                
                all_evaluations.append(evaluation)
                
                # Add delay to respect API rate limits
                time.sleep(1)
        
        # Save results to Excel
        self.save_evaluations_to_excel(all_evaluations, output_file)
        
        print(f"\n🎉 Batch evaluation completed!")
        print(f"📊 Total evaluations: {len(all_evaluations)}")
        print(f"💾 Results saved to: {output_file}")
        
        return all_evaluations
    
    def save_evaluations_to_excel(self, evaluations: List[MiniMaxEvaluation], output_file: str):
        """Save evaluations to Excel file"""
        try:
            # Convert evaluations to DataFrame
            data = []
            for eval in evaluations:
                data.append({
                    'Case_ID': eval.case_id,
                    'Model_Name': eval.model_name,
                    'Patient_Text': eval.patient_text,
                    'Treatment_Plan': eval.treatment_plan,
                    'Rating': eval.rating,
                    'Score_Justification': eval.score_justification,
                    'Key_Observations': eval.key_observations,
                    'Overall_Assessment': eval.overall_assessment,
                    'Clinical_Reasoning': eval.clinical_reasoning,
                    'Safety_Considerations': eval.safety_considerations,
                    'Evidence_Quality': eval.evidence_quality,
                    'Implementation_Feasibility': eval.implementation_feasibility,
                    'Specialist_Notes': eval.specialist_notes,
                    'Thinking_Process': eval.thinking_process,
                    'Evaluation_Timestamp': eval.evaluation_timestamp
                })
            
            df = pd.DataFrame(data)
            df.to_excel(output_file, index=False, sheet_name='MiniMax_Evaluations')
            print(f"✅ Results saved to Excel: {output_file}")
            
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")
            # Fallback: save as JSON
            json_file = output_file.replace('.xlsx', '.json')
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump([asdict(eval) for eval in evaluations], f, indent=2, ensure_ascii=False)
            print(f"💾 Fallback: Results saved as JSON: {json_file}")
    
    def generate_summary_report(self, evaluations: List[MiniMaxEvaluation]) -> str:
        """Generate summary statistics and analysis"""
        if not evaluations:
            return "No evaluations to analyze."
        
        # Calculate statistics
        ratings = [eval.rating for eval in evaluations if eval.rating]
        rating_counts = {}
        for rating in ratings:
            rating_counts[rating] = rating_counts.get(rating, 0) + 1
        
        # Group by model
        model_stats = {}
        for eval in evaluations:
            model = eval.model_name
            if model not in model_stats:
                model_stats[model] = {
                    'count': 0,
                    'ratings': []
                }
            model_stats[model]['count'] += 1
            model_stats[model]['ratings'].append(eval.rating)
        
        # Calculate A+B percentages
        ab_percentages = {}
        for model, stats in model_stats.items():
            model_ratings = stats['ratings']
            ab_count = model_ratings.count('A') + model_ratings.count('B')
            ab_percentages[model] = (ab_count / len(model_ratings)) * 100 if model_ratings else 0
        
        # Generate report
        report = f"""# MiniMax Simplified Clinical Evaluation Summary Report
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Overview
- Total Evaluations: {len(evaluations)}
- Models Evaluated: {len(model_stats)}
- Cases Processed: {len(set(eval.case_id for eval in evaluations))}

## Overall Rating Distribution
"""
        for rating in ['A', 'B', 'C', 'D', 'E']:
            count = rating_counts.get(rating, 0)
            percentage = (count / len(evaluations)) * 100 if evaluations else 0
            report += f"- {rating}: {count} ({percentage:.1f}%)\n"
        
        report += "\n## Model Performance Analysis\n"
        for model, stats in model_stats.items():
            model_ratings = stats['ratings']
            ab_rate = ab_percentages[model]
            report += f"\n### {model}\n"
            report += f"- Evaluations: {stats['count']}\n"
            report += f"- A+B Rate: {ab_rate:.1f}%\n"
            
            for rating in ['A', 'B', 'C', 'D', 'E']:
                count = model_ratings.count(rating)
                percentage = (count / len(model_ratings)) * 100 if model_ratings else 0
                report += f"- {rating}: {count} ({percentage:.1f}%)\n"
        
        # Add performance ranking
        ranked_models = sorted(model_stats.keys(), key=lambda m: ab_percentages[m], reverse=True)
        report += "\n## Performance Ranking (by A+B Rate)\n"
        for i, model in enumerate(ranked_models, 1):
            report += f"{i}. {model}: {ab_percentages[model]:.1f}% A+B rate\n"
        
        return report

def main():
    """Main function to demonstrate usage"""
    print("🏥 Simplified MiniMax Clinical Evaluation System")
    print("=" * 55)
    
    # Initialize evaluator
    evaluator = SimplifiedMiniMaxEvaluator()
    
    # Get API key
    api_key = input("eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJHcm91cE5hbWUiOiJBbGkgQXJzaGFkIiwiVXNlck5hbWUiOiJBbGkgQXJzaGFkIiwiQWNjb3VudCI6IiIsIlN1YmplY3RJRCI6IjE5ODUwMzYyMTQ1MDc0MTQ0NTAiLCJQaG9uZSI6IiIsIkdyb3VwSUQiOiIxOTg1MDM2MjE0NDk5MDIxNzQ2IiwiUGFnZU5hbWUiOiIiLCJNYWlsIjoiYWxpYXJzaGFkY2lpdHN3bEBnbWFpbC5jb20iLCJDcmVhdGVUaW1lIjoiMjAyNS0xMS0wMyAwMzoyNDo0MyIsIlRva2VuVHlwZSI6MSwiaXNzIjoibWluaW1heCJ9.H6DmdOBYPfWcIGTszfoZrJ6Zoq24AsyGD7fsAhtQsUne5pLoKWhQPJHrGGwTnH2RLZ5x6wRZfWiCsZMJfRQyxzXqfw4cp7iCljxIWZHsAJV8SSN7majhZ785YAe8WYP_zYiwCVPNWO5VR1K8JSm3owWZrVAAhNaiGvqEBFXFjtH_gfblXVPg_yZ2UHqWs0j4aAjt1DlFWXjlwlJ8vpFuGmlusZju4ZmD107otasJPKf9JrtPAmSS50ptcuIxLuVJNoRPP4cxc2u48tltcuNr7xS89wCEjU60PMaIf66zMGyHUsyKUbLCPHj5XfJWq-6rvYLIaomWeEHzLlGU3EanKQ").strip()
    if not api_key:
        print("❌ API key is required to proceed")
        return
    
    # Set API key
    evaluator.set_api_key(api_key)
    
    # Get file paths
    excel_file = "D:\Sleep Medicine\Book1.xlsx"
    output_file = input("Output filename (default: minimax_simplified_results.xlsx): ").strip()
    if not output_file:
        output_file = "minimax_simplified_results.xlsx"
    
    print(f"\n📁 Input file: {excel_file}")
    print(f"💾 Output file: {output_file}")
    print(f"📊 Excel structure: Patient Text + 5 Treatment Plans")
    
    # Get evaluation range
    start_row = input("Start row (default 3): ").strip()
    start_row = int(start_row) if start_row else 3
    
    end_row = input("End row (leave empty for all): ").strip()
    end_row = int(end_row) if end_row else None
    
    # Confirm
    print(f"\n🚀 Ready to evaluate:")
    print(f"   - Starting from row {start_row}")
    if end_row:
        print(f"   - Ending at row {end_row}")
    else:
        print(f"   - Ending at last row")
    print(f"   - Using MiniMax-M2 model")
    print(f"   - Simplified Excel structure")
    
    confirm = input(f"\nProceed? (y/n): ").strip().lower()
    if confirm != 'y':
        print("❌ Evaluation cancelled")
        return
    
    # Run evaluation
    try:
        print(f"\n🎯 Starting evaluation...")
        evaluations = evaluator.batch_evaluate_simplified_excel(
            excel_file_path=excel_file,
            output_file=output_file,
            start_row=start_row,
            end_row=end_row
        )
        
        # Generate summary
        print(f"\n📋 Generating summary report...")
        summary = evaluator.generate_summary_report(evaluations)
        summary_file = output_file.replace('.xlsx', '_summary.md')
        
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write(summary)
        
        print(f"\n🎉 Evaluation completed successfully!")
        print(f"📊 Total evaluations: {len(evaluations)}")
        print(f"💾 Detailed results: {output_file}")
        print(f"📋 Summary report: {summary_file}")
        
        # Show quick stats
        ratings = [eval.rating for eval in evaluations if eval.rating]
        if ratings:
            print(f"\n📈 Quick Stats:")
            for rating in ['A', 'B', 'C', 'D', 'E']:
                count = ratings.count(rating)
                percentage = (count / len(ratings)) * 100 if ratings else 0
                print(f"   {rating}: {count} ({percentage:.1f}%)")
        
    except Exception as e:
        print(f"❌ Error during evaluation: {e}")
        print(f"💡 Please check your API key and internet connection")

if __name__ == "__main__":
    main()
