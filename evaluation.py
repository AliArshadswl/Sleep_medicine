#!/usr/bin/env python3
"""
Treatment Plan Evaluation System for Clinical Cases
Focused on evaluating only treatment recommendations using A-E rating scale
"""

import pandas as pd
import json
import re
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl import load_workbook

class TreatmentPlanEvaluator:
    def __init__(self):
        self.evaluation_criteria = {
            'A': {
                'name': 'Excellent',
                'description': 'Outstanding recommendation, highly clinically reliable with clear reasoning, appropriate risk awareness, and practical implementation',
                'score_range': (4.5, 5.0)
            },
            'B': {
                'name': 'Very Good', 
                'description': 'Strong recommendation with good reasoning and justifications, minor areas for improvement',
                'score_range': (3.5, 4.4)
            },
            'C': {
                'name': 'Good',
                'description': 'Solid recommendation that\'s mostly accurate and reasonable, some general advice could be more specific', 
                'score_range': (2.5, 3.4)
            },
            'D': {
                'name': 'Fair',
                'description': 'Basic recommendation with some correct elements but notable gaps in safety, specificity, or clinical appropriateness',
                'score_range': (1.5, 2.4)
            },
            'E': {
                'name': 'Poor', 
                'description': 'Inadequate recommendation with significant safety concerns, inappropriate suggestions, or lack of clinical relevance',
                'score_range': (1.0, 1.4)
            }
        }
    
    def load_excel_data(self, file_path: str) -> List[Dict[str, Any]]:
        """Load treatment plan data from Excel file"""
        try:
            # Load the workbook
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            # Find treatment plan columns
            treatment_columns = self._find_treatment_columns(headers)
            
            # Load all data
            cases = []
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                
                # Extract case information
                case_info = self._extract_case_info(row_data, headers)
                if case_info and case_info.get('patient_text'):
                    # Extract treatment plans
                    treatment_plans = self._extract_treatment_plans_from_row(row_data, treatment_columns)
                    if treatment_plans:
                        case_info['treatment_plans'] = treatment_plans
                        cases.append(case_info)
            
            return cases
            
        except Exception as e:
            print(f"Error loading Excel data: {e}")
            return []
    
    def _find_treatment_columns(self, headers: List[str]) -> Dict[str, str]:
        """Find treatment plan columns in the data"""
        treatment_columns = {}
        
        for i, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                
                # Look for treatment plan columns
                if 'treatment' in header_lower or 'management' in header_lower:
                    if 'claude' in header_lower:
                        treatment_columns['Claude'] = header
                    elif 'deepseek' in header_lower:
                        treatment_columns['Deepseek'] = header
                    elif 'gpt' in header_lower and 'grok' not in header_lower:
                        treatment_columns['ChatGPT'] = header
                    elif 'grok' in header_lower:
                        treatment_columns['Grok'] = header
                    elif 'gpt-5' in header_lower or 'gpt5' in header_lower:
                        treatment_columns['GPT-5'] = header
        
        return treatment_columns
    
    def _extract_case_info(self, row_data: Dict, headers: List[str]) -> Optional[Dict[str, Any]]:
        """Extract case information from row data"""
        case_info = {}
        
        # Common case information fields
        case_mappings = {
            'patient text': 'patient_text',
            'diagnosis': 'diagnosis', 
            'clinical specialty': 'specialty',
            'book': 'book',
            'publisher': 'publisher'
        }
        
        for header in headers:
            if header and isinstance(header, str):
                header_lower = header.lower()
                
                # Check for patient text
                if 'patient' in header_lower and 'text' in header_lower:
                    case_info['patient_text'] = str(row_data.get(header, ''))
                elif 'diagnosis' in header_lower:
                    case_info['diagnosis'] = str(row_data.get(header, ''))
                elif 'clinical specialty' in header_lower:
                    case_info['specialty'] = str(row_data.get(header, ''))
                elif 'book' in header_lower:
                    case_info['book'] = str(row_data.get(header, ''))
                elif 'publisher' in header_lower:
                    case_info['publisher'] = str(row_data.get(header, ''))
        
        return case_info if case_info.get('patient_text') else None
    
    def _extract_treatment_plans_from_row(self, row_data: Dict, treatment_columns: Dict[str, str]) -> Dict[str, str]:
        """Extract treatment plans from row data"""
        treatment_plans = {}
        
        for model_name, column_name in treatment_columns.items():
            treatment_text = str(row_data.get(column_name, ''))
            if treatment_text and treatment_text != 'nan' and len(treatment_text.strip()) > 50:
                treatment_plans[model_name] = treatment_text.strip()
        
        return treatment_plans
    
    def evaluate_treatment_plan(self, patient_case: str, treatment_plan: str, model_name: str) -> Dict[str, Any]:
        """Evaluate a treatment plan using clinical criteria"""
        
        if not treatment_plan or len(treatment_plan.strip()) < 50:
            return self._empty_evaluation(model_name, "Insufficient treatment plan content")
        
        # Score each component (0-5 scale)
        scores = {
            'clinical_appropriateness': self._score_clinical_appropriateness(treatment_plan, patient_case),
            'safety_awareness': self._score_safety_awareness(treatment_plan),
            'clarity_structure': self._score_clarity_structure(treatment_plan),
            'practical_feasibility': self._score_practical_feasibility(treatment_plan),
            'evidence_based_reasoning': self._score_evidence_reasoning(treatment_plan)
        }
        
        # Calculate overall score
        overall_score = sum(scores.values()) / len(scores)
        
        # Convert to A-E rating
        rating = self._score_to_rating(overall_score)
        
        return {
            'model': model_name,
            'rating': rating,
            'overall_score': round(overall_score, 2),
            'component_scores': {k: round(v, 2) for k, v in scores.items()},
            'justification': self._generate_justification(scores, rating),
            'key_observations': self._generate_key_observations(scores, treatment_plan),
            'overall_assessment': self._generate_overall_assessment(rating, treatment_plan),
            'treatment_plan_excerpt': treatment_plan[:300] + "..." if len(treatment_plan) > 300 else treatment_plan
        }
    
    def _empty_evaluation(self, model_name: str, reason: str) -> Dict[str, Any]:
        """Return empty evaluation for missing/invalid treatment plans"""
        return {
            'model': model_name,
            'rating': 'E',
            'overall_score': 0.0,
            'component_scores': {
                'clinical_appropriateness': 0.0,
                'safety_awareness': 0.0,
                'clarity_structure': 0.0,
                'practical_feasibility': 0.0,
                'evidence_based_reasoning': 0.0
            },
            'justification': f'Inadequate recommendation: {reason}',
            'key_observations': 'Incomplete or missing treatment plan',
            'overall_assessment': 'Inadequate - Insufficient content for evaluation',
            'treatment_plan_excerpt': 'No treatment plan provided'
        }
    
    def _score_clinical_appropriateness(self, treatment: str, case: str) -> float:
        """Score clinical appropriateness (0-5)"""
        treatment_lower = treatment.lower()
        case_lower = case.lower() if case else ""
        
        score = 0.0
        
        # Medication appropriateness
        if any(med in treatment_lower for med in ['carbamazepine', 'oxcarbazepine', 'antiepileptic']):
            score += 1.5
        if any(med in treatment_lower for med in ['clonazepam', 'benzodiazepine']):
            score += 1.0
        if any(med in treatment_lower for med in ['melatonin', 'gabapentin']):
            score += 0.8
        
        # Non-pharmacological approaches
        if 'sleep hygiene' in treatment_lower:
            score += 1.0
        if any(term in treatment_lower for term in ['therapy', 'counseling', 'education']):
            score += 0.8
        
        # Condition-specific treatments
        if 'epilepsy' in case_lower and any(term in treatment_lower for term in ['carbamazepine', 'antiepileptic']):
            score += 1.0
        elif 'parasomnia' in case_lower and 'sleep hygiene' in treatment_lower:
            score += 1.0
        elif 'dyskinesia' in case_lower and any(term in treatment_lower for term in ['carbamazepine', 'clonazepam']):
            score += 1.0
        
        return min(score, 5.0)
    
    def _score_safety_awareness(self, treatment: str) -> float:
        """Score safety awareness (0-5)"""
        treatment_lower = treatment.lower()
        score = 0.0
        
        # Safety monitoring
        if any(term in treatment_lower for term in ['monitor', 'monitoring', 'follow-up']):
            score += 1.0
        
        # Side effects and contraindications
        if any(term in treatment_lower for term in ['side effect', 'contraindication', 'adverse']):
            score += 1.2
        
        # Risk assessment
        if any(term in treatment_lower for term in ['risk', 'safety', 'precaution']):
            score += 0.8
        
        # Dosage considerations
        if any(term in treatment_lower for term in ['dose', 'dosage', 'titrate', 'start low']):
            score += 1.0
        
        # Specific safety terms
        if any(term in treatment_lower for term in ['hyponatremia', 'sedation', 'dizziness']):
            score += 0.8
        
        return min(score, 5.0)
    
    def _score_clarity_structure(self, treatment: str) -> float:
        """Score clarity and structure (0-5)"""
        treatment_lower = treatment.lower()
        score = 0.0
        
        # Structured format
        if any(pattern in treatment_lower for pattern in ['1.', '2.', '3.', '4.', '5.']):
            score += 1.5
        
        # Clear sections
        if any(term in treatment_lower for term in ['purpose:', 'details:', 'note:', 'goal:']):
            score += 1.0
        
        # Sequential steps
        if any(term in treatment_lower for term in ['first', 'second', 'third', 'then', 'next']):
            score += 0.8
        
        # Explanations
        if any(term in treatment_lower for term in ['because', 'since', 'due to', 'as']):
            score += 0.8
        
        # Professional language
        if any(term in treatment_lower for term in ['recommend', 'suggest', 'advise', 'consider']):
            score += 0.9
        
        return min(score, 5.0)
    
    def _score_practical_feasibility(self, treatment: str) -> float:
        """Score practical feasibility (0-5)"""
        treatment_lower = treatment.lower()
        score = 0.0
        
        # Scheduling and timing
        if any(term in treatment_lower for term in ['schedule', 'timing', 'frequency', 'regular']):
            score += 1.0
        
        # Dosage and administration
        if any(term in treatment_lower for term in ['dose', 'dosage', 'mg', 'bedtime', 'evening']):
            score += 1.2
        
        # Follow-up and monitoring
        if any(term in treatment_lower for term in ['follow-up', 'appointment', 'visit']):
            score += 0.8
        
        # Lifestyle modifications
        if any(term in treatment_lower for term in ['lifestyle', 'diet', 'exercise', 'avoid']):
            score += 0.8
        
        # Patient diary/log
        if any(term in treatment_lower for term in ['diary', 'log', 'track', 'record']):
            score += 0.7
        
        return min(score, 5.0)
    
    def _score_evidence_reasoning(self, treatment: str) -> float:
        """Score evidence-based reasoning (0-5)"""
        treatment_lower = treatment.lower()
        score = 0.0
        
        # Evidence-based terms
        if any(term in treatment_lower for term in ['evidence', 'guideline', 'recommend', 'standard']):
            score += 1.2
        
        # First-line treatments
        if any(term in treatment_lower for term in ['first-line', 'first line', 'initial']):
            score += 1.0
        
        # Clinical rationale
        if any(term in treatment_lower for term in ['clinical', 'effective', 'efficacy']):
            score += 0.8
        
        # Treatment options
        if any(term in treatment_lower for term in ['alternative', 'second-line', 'option']):
            score += 0.8
        
        # Professional consensus
        if any(term in treatment_lower for term in ['commonly', 'typically', 'usually']):
            score += 0.7
        
        return min(score, 5.0)
    
    def _score_to_rating(self, score: float) -> str:
        """Convert numerical score to A-E rating"""
        if score >= 4.5:
            return 'A'
        elif score >= 3.5:
            return 'B'
        elif score >= 2.5:
            return 'C'
        elif score >= 1.5:
            return 'D'
        else:
            return 'E'
    
    def _generate_justification(self, scores: Dict[str, float], rating: str) -> str:
        """Generate rating justification"""
        rating_info = self.evaluation_criteria[rating]
        
        strengths = []
        if scores['clinical_appropriateness'] >= 3.5:
            strengths.append("excellent clinical appropriateness")
        elif scores['clinical_appropriateness'] >= 2.5:
            strengths.append("good clinical approach")
        
        if scores['safety_awareness'] >= 3.5:
            strengths.append("strong safety considerations")
        elif scores['safety_awareness'] >= 2.5:
            strengths.append("adequate safety measures")
        
        if scores['clarity_structure'] >= 3.5:
            strengths.append("well-organized presentation")
        elif scores['clarity_structure'] >= 2.5:
            strengths.append("clear structure")
        
        if scores['practical_feasibility'] >= 3.5:
            strengths.append("highly practical approach")
        elif scores['practical_feasibility'] >= 2.5:
            strengths.append("feasible implementation")
        
        if scores['evidence_based_reasoning'] >= 3.5:
            strengths.append("strong evidence base")
        elif scores['evidence_based_reasoning'] >= 2.5:
            strengths.append("adequate clinical rationale")
        
        if not strengths:
            strengths.append("basic clinical knowledge")
        
        return f"{rating_info['description']}. Key strengths: {', '.join(strengths[:3])}."
    
    def _generate_key_observations(self, scores: Dict[str, float], treatment: str) -> str:
        """Generate key observations"""
        observations = []
        
        if scores['clinical_appropriateness'] < 2.0:
            observations.append("Limited clinical specificity for case")
        if scores['safety_awareness'] < 2.0:
            observations.append("Insufficient safety monitoring")
        if scores['clarity_structure'] < 2.0:
            observations.append("Needs better organization")
        if scores['practical_feasibility'] < 2.0:
            observations.append("Lacks practical implementation details")
        if scores['evidence_based_reasoning'] < 2.0:
            observations.append("Limited evidence-based justification")
        
        if not observations:
            observations.append("Demonstrates comprehensive clinical reasoning")
        
        return "; ".join(observations)
    
    def _generate_overall_assessment(self, rating: str, treatment: str) -> str:
        """Generate overall assessment"""
        if rating in ['A', 'B']:
            return f"High-quality clinical recommendation suitable for healthcare provider consideration in managing this patient's condition."
        elif rating == 'C':
            return f"Acceptable clinical recommendation that provides useful guidance but could benefit from more specificity and detail."
        elif rating == 'D':
            return f"Basic recommendation with some clinical value but requires significant enhancement for clinical application."
        else:
            return f"Inadequate recommendation that does not meet clinical standards and requires substantial revision."
    
    def process_all_treatments(self, file_path: str) -> List[Dict[str, Any]]:
        """Process all treatment plans from the Excel file"""
        cases = self.load_excel_data(file_path)
        results = []
        
        for i, case in enumerate(cases, 1):
            case_result = {
                'case_number': i,
                'case_info': {
                    'diagnosis': case.get('diagnosis', 'Unknown'),
                    'specialty': case.get('specialty', 'Unknown'),
                    'patient_description': case['patient_text'][:200] + "..." if len(case['patient_text']) > 200 else case['patient_text']
                },
                'treatment_evaluations': []
            }
            
            # Evaluate each treatment plan
            for model_name, treatment_plan in case['treatment_plans'].items():
                evaluation = self.evaluate_treatment_plan(
                    case['patient_text'],
                    treatment_plan,
                    model_name
                )
                case_result['treatment_evaluations'].append(evaluation)
            
            results.append(case_result)
        
        return results
    
    def save_results(self, results: List[Dict[str, Any]], output_file: str):
        """Save results to JSON file"""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
    
    def generate_report(self, results: List[Dict[str, Any]]) -> str:
        """Generate comprehensive evaluation report"""
        report_lines = []
        report_lines.append("# Clinical Treatment Plan Evaluation Report")
        report_lines.append("=" * 50)
        report_lines.append("")
        
        # Summary statistics
        total_cases = len(results)
        total_evaluations = sum(len(case['treatment_evaluations']) for case in results)
        
        report_lines.append(f"## Executive Summary")
        report_lines.append(f"- **Total Cases Evaluated**: {total_cases}")
        report_lines.append(f"- **Total Treatment Plans Evaluated**: {total_evaluations}")
        report_lines.append("")
        
        # Model performance summary
        model_stats = {}
        for case in results:
            for evaluation in case['treatment_evaluations']:
                model = evaluation['model']
                if model not in model_stats:
                    model_stats[model] = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'total_score': 0, 'count': 0}
                
                model_stats[model][evaluation['rating']] += 1
                model_stats[model]['total_score'] += evaluation['overall_score']
                model_stats[model]['count'] += 1
        
        report_lines.append("## Model Performance Summary")
        for model, stats in model_stats.items():
            avg_score = stats['total_score'] / stats['count'] if stats['count'] > 0 else 0
            report_lines.append(f"### {model}")
            report_lines.append(f"- **Average Score**: {avg_score:.2f}/5.0")
            for grade in ['A', 'B', 'C', 'D', 'E']:
                count = stats[grade]
                percentage = (count / stats['count'] * 100) if stats['count'] > 0 else 0
                report_lines.append(f"- {grade}: {count} ({percentage:.1f}%)")
            report_lines.append("")
        
        # Detailed case results
        report_lines.append("## Detailed Case Evaluations")
        report_lines.append("")
        
        for case in results:
            report_lines.append(f"### Case {case['case_number']}: {case['case_info']['diagnosis']}")
            report_lines.append(f"**Specialty**: {case['case_info']['specialty']}")
            report_lines.append(f"**Patient**: {case['case_info']['patient_description']}")
            report_lines.append("")
            
            for evaluation in case['treatment_evaluations']:
                report_lines.append(f"#### {evaluation['model']} Treatment Plan")
                report_lines.append(f"**Rating**: {evaluation['rating']} ({evaluation['overall_score']}/5.0)")
                report_lines.append(f"**Component Scores**:")
                for component, score in evaluation['component_scores'].items():
                    report_lines.append(f"  - {component.replace('_', ' ').title()}: {score}/5.0")
                report_lines.append(f"**Justification**: {evaluation['justification']}")
                report_lines.append(f"**Key Observations**: {evaluation['key_observations']}")
                report_lines.append(f"**Assessment**: {evaluation['overall_assessment']}")
                report_lines.append("")
        
        return "\n".join(report_lines)

def main():
    """Main function to run treatment plan evaluation"""
    evaluator = TreatmentPlanEvaluator()
    
    # File path
    input_file = "/workspace/user_input_files/Copy of Copy of Cases_to_evaluate (1) - Final_(1).xlsx"
    
    print("Loading and evaluating treatment plans...")
    results = evaluator.process_all_treatments(input_file)
    
    print(f"Successfully processed {len(results)} cases")
    
    # Save results
    output_file = "/workspace/treatment_evaluations.json"
    evaluator.save_results(results, output_file)
    print(f"Results saved to: {output_file}")
    
    # Generate report
    report = evaluator.generate_report(results)
    report_file = "/workspace/treatment_evaluation_report.md"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"Report saved to: {report_file}")
    
    # Print summary
    print("\n" + "="*50)
    print("EVALUATION SUMMARY")
    print("="*50)
    
    for case in results:
        print(f"\nCase {case['case_number']}: {case['case_info']['diagnosis']}")
        for evaluation in case['treatment_evaluations']:
            print(f"  {evaluation['model']}: {evaluation['rating']} ({evaluation['overall_score']}/5.0)")

if __name__ == "__main__":
    main()
